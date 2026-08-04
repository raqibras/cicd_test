"""
Aggregates Azure cost-export CSV rows by Resource and builds the
consolidated Excel report.
"""

import csv
from io import BytesIO
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from config import (
    RESOURCE_GROUP_TO_PROJECT_MAPPING,
    RESOURCE_TO_PROJECT_MAPPING,
    INCLUDED_RESOURCE_GROUPS,
)

UNMAPPED_LABEL = "Unmapped"

REPORT_COLUMNS = [
    "SI No.",
    "Resource Name",
    "Resource Group",
    "Service",
    "Application",
    "Subscription",
    "Location",
    "Cost (USD)",
]


def _resolve_application(resource_name: str, resource_group: str) -> str:
    """Resource-level mapping wins; resource-group mapping is the fallback."""
    if resource_name in RESOURCE_TO_PROJECT_MAPPING:
        return RESOURCE_TO_PROJECT_MAPPING[resource_name]
    if resource_group in RESOURCE_GROUP_TO_PROJECT_MAPPING:
        return RESOURCE_GROUP_TO_PROJECT_MAPPING[resource_group]
    return UNMAPPED_LABEL


def _parse_cost(raw_value) -> float:
    try:
        return float(raw_value) if raw_value not in (None, "") else 0.0
    except ValueError:
        return 0.0


def aggregate_resource_costs(
    csv_reader: csv.DictReader,
    filter_by_included_groups: bool = True,
) -> list[dict]:
    """
    Aggregates CostUSD by Resource, carrying along the metadata columns
    needed for the consolidated report (resource group, service/type,
    application, subscription, location).

    Rows sharing the same Resource name are summed. If a resource somehow
    shows up under more than one ResourceGroupName/Subscription in the
    export, the first value seen is kept for those descriptive fields
    (only the cost is accumulated).
    """
    resources: dict[str, dict] = {}

    for row in csv_reader:
        resource_name = (row.get("Resource") or "").strip()
        if not resource_name:
            continue

        resource_group = (row.get("ResourceGroupName") or "").strip()

        if filter_by_included_groups and INCLUDED_RESOURCE_GROUPS:
            if resource_group not in INCLUDED_RESOURCE_GROUPS:
                continue

        cost = _parse_cost(row.get("CostUSD") or row.get("Cost"))

        if resource_name not in resources:
            resources[resource_name] = {
                "resource_name": resource_name,
                "resource_group": resource_group,
                "service": (row.get("ResourceType") or "").strip(),
                "application": _resolve_application(resource_name, resource_group),
                "subscription": (row.get("SubscriptionName") or "").strip(),
                "location": (row.get("ResourceLocation") or "").strip(),
                "cost": 0.0,
            }

        resources[resource_name]["cost"] += cost

    # Highest cost first
    return sorted(resources.values(), key=lambda r: r["cost"], reverse=True)


def build_excel_report(resources: Iterable[dict]) -> BytesIO:
    """Builds the consolidated .xlsx report and returns it as an in-memory buffer."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Cost Report"

    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    body_font = Font(name="Arial", size=10)
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    # Header row
    for col_idx, header in enumerate(REPORT_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Data rows
    row_idx = 2
    total_cost = 0.0
    for si_no, resource in enumerate(resources, start=1):
        values = [
            si_no,
            resource["resource_name"],
            resource["resource_group"],
            resource["service"],
            resource["application"],
            resource["subscription"],
            resource["location"],
            round(resource["cost"], 2),
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = body_font
            cell.border = thin_border
            if col_idx == 1 or col_idx == 8:
                cell.alignment = Alignment(horizontal="center")
        ws.cell(row=row_idx, column=8).number_format = "$#,##0.00"
        total_cost += resource["cost"]
        row_idx += 1

    # Totals row
    total_row = row_idx
    ws.cell(row=total_row, column=7, value="Total").font = Font(name="Arial", bold=True, size=10)
    ws.cell(row=total_row, column=7).alignment = Alignment(horizontal="right")
    total_cell = ws.cell(row=total_row, column=8, value=round(total_cost, 2))
    total_cell.font = Font(name="Arial", bold=True, size=10)
    total_cell.number_format = "$#,##0.00"
    total_cell.alignment = Alignment(horizontal="center")
    for col_idx in range(1, 9):
        ws.cell(row=total_row, column=col_idx).border = thin_border

    # Column widths
    widths = [8, 30, 22, 26, 22, 24, 16, 16]
    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:H{max(total_row - 1, 1)}"

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer